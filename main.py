import sys
import warnings
import re
import pyodbc
from openpyxl import Workbook, worksheet, load_workbook
from configparser import ConfigParser

conf = ConfigParser()
conf.read("config.ini")

DB_CONFIG = conf["DB"]
DB_DRIVER = DB_CONFIG["driver"]
DB_PATH = DB_CONFIG["dbpath"]

MIN_ROW: int = 3
MIN_COL: int = 1

DRIVER = f"Driver={DB_DRIVER}"
DQB = f"DBQ={DB_PATH}"

warnings.filterwarnings("ignore", category = UserWarning, module = "openpyxl")

def usage():
    msg = "Usage: python3 main.py <file_name>"
    print(msg)

def extract_timesheet(sheet: worksheet) -> dict:
    timesheet = {}
    for row in sheet.iter_rows( min_row = MIN_ROW,
                                min_col = MIN_COL,
                                max_row = sheet.max_row,
                                max_col = sheet.max_column):
        commessa = row[0].value
        ore = []

        if commessa not in timesheet:
            timesheet[commessa] = []

        for r in row[2:]:
            ore.append(r.value)

        timesheet[commessa].extend(ore)

    return timesheet

def extract_timesheets(wb: Workbook, sheets: list[str]) -> list:
    timesheets = []

    for sheet in sheets:
        ts = extract_timesheet(wb[sheet])

        for commessa in ts.keys():
            s = sum( [ float(ore) for ore in ts[commessa] if ore is not None ] )
            if s != 0:
                timesheets.append( [sheet, commessa, s] )

    return timesheets

def main():
    if len(sys.argv) != 2:
        usage()
        exit(1)

    file_name: str = sys.argv[1]

    wb = load_workbook(filename = file_name, read_only = True)
    sheets = wb.sheetnames
    if "QB_COMMESSE" in sheets: sheets.remove("QB_COMMESSE")
    if "Chiusura Mese" in sheets: sheets.remove("Chiusura Mese")
    periodo = re.search("[0-9]{6}", file_name).group(0)

    ts = extract_timesheets(wb, sheets)
    conn = pyodbc.connect(f"{DRIVER}; {DQB};")
    cursor = conn.cursor()

    for t in ts:
        dipendente = t[0]
        commessa = t[1]
        ore = t[2]

        query = """
            insert into rendicontazione
                    (Periodo, Dipendente, Commessa, Ore)
            values  (?, ?, ?, ?)
        """
        cursor.execute(query, periodo, dipendente, commessa, ore)
    
    cursor.commit()
    conn.close()

main()